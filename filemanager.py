from datetime import date, timedelta
import settingsMenu as settings
import os
import openpyxl

monthTable = [
	"null",
	"Jan",
	"Feb",
	"Mar",
	"Apr",
	"May",
	"Jun",
	"Jul",
	"Aug",
	"Sep",
	"Oct",
	"Nov",
	"Dec"
]

def GetFileName(day : date = date.today()):
	"""given a date, this will return a string in the form of:\n
		(first three letters of the month) (day)\n
	   to be used as the file name.\n
	   Example: "Nov 27"
	"""
	name = monthTable[day.month] + " " + str(day.day)

	return name
#END

def GetTodaysFileWithPath():
	return settings.GetSetting("savepath") + GetFileName() + ".xlsx"

def GetDateFromFileName(name : str):
	"""Opposite operation of "GetFileName"\n
		it takes a file name and returns a date object, or None if the file isn't in the correct format
	"""

	from os.path import basename
	name = basename(name) #remove path and keep filename only

	month = name[0] + name[1] + name[2] #first three letters are the month
	for m in range(13): #dumb but the table is only 13 elements long so it's fine
		if monthTable[m] == month: #if we find the month we set the number to the current index and break
			month = m
			break
	#END FOR

	day = 0
	try:
		try: int(name[4])
		#if this happens there is no valid day number
		except ValueError: return None

		digit1 = int(name[4])

		#this could give an IndexError or ValueError, so we catch it down
		digit2 = int(name[5])

		#if we don't get errors we can convert like this
		day = digit1 * 10 + digit2
	except:
		#we reach this if name[5] gives an outofrange
		day = digit1
	
	#if the month is higher than current month than we must be an year ahead, so subtract 1
	year = date.today().year - (0 if month <= date.today().month else 1)

	try: return date(year, month, day)
	except: return None #if there's an error here an invalid date is entered
#END

def DeleteOldFiles(olderThanDays : int):
	"""
	Deletes all files in the save directory that are at least 'olderThanDays' days older than the current date\n
		Parameter olderThanDays: an integer showing the number of days old the file must be for it to be deleted
	"""

	from glob import glob #find all files
	from os import remove #delete file
	directory = settings.GetSetting("savepath")  #directory for saved files
	extension = "*.xlsx"  #extension of the files (* means "any file ending with the following")

	#any file with this date or earlier will be deleted
	mindate = date.today() - timedelta(days=olderThanDays)

	print(mindate)

	#get all files in the directory
	for file in glob(directory + extension):
		filedate = GetDateFromFileName(file) #get the date the file represents
		#if the date is null the file is something else or something marked as don't delete so we skip it
		print(filedate)
		if filedate is None: continue
		if filedate <= mindate: #check if the file is older than the date
			remove(file)  #if it is delete it
#END

def SaveSetting(name, value):
	f = openpyxl.load_workbook("settings.xlsx")
	file = f.active

	if name == "deleteTime":
		file["A1"] = int(value)
	elif name == "savepath":
		if not value[-1] == "/":
			value += "/"
		file["A2"] = value
	elif name == "float":
		file["B1"] = float(value)
	elif name == "theme":
		file["B2"] = value
	elif name == "unit_temp":
		file["C1"] = value
	else: raise ValueError("'" + name + "' is not a valid setting name")
	
	f.save("settings.xlsx")
	f.close()




def ReadSetting(name):
	f = openpyxl.load_workbook("settings.xlsx")
	file = f.active
	
	if name == "deleteTime":
		retval = file["A1"]
	elif name == "savepath":
		retval = file["A2"]
	elif name == "float":
		retval = file["B1"]
	elif name == "theme":
		retval = file["B2"]
	elif name == "unit_temp":
		retval = file["C1"]
	else: raise ValueError("'" + name + "' is not a valid setting name")

	f.close()
	return retval.value

def FileForTodayExists():
	return os.path.isfile(GetTodaysFileWithPath())

#REGION INIT

#test if the file is not corrupted. if it is, delete it and create a new one
#luckily we won't need to add new settings, it would be a nightmare lmao (though we could create a "settings" class and an array with all settings and have a loop but we don't need to get that fancy)
def TestSettings():
	try:
		#try to see if the file exists
		f = openpyxl.open("settings.xlsx")
		f.close()

		if not 0 < int(ReadSetting("deleteTime")) <= 365:
			raise
		
		float(ReadSetting("float"))

		if not os.path.isdir(ReadSetting("savepath")):
			raise
		
		int(ReadSetting("theme"))

	except:

		print("ERROR: the settings file is either not found or corrupted")

		#if the file does not exist remove will give an exception want an exception
		try:
			os.remove("settings.xlsx")
		except:pass

		#create a setting file
		wb = openpyxl.Workbook("settings.xlsx")
		wb.save("settings.xlsx")
		settings.SetToDefault()
#END


TestSettings()

DeleteOldFiles(settings.GetSetting("deleteTime"))

#ENDREGION INIT